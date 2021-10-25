"""
功能：将1w多台设备一个月的cpu品均值，cpu最大值，内存平均值，内存最大值整合算平均

将多表作为sheet合入一个表格
使用ExcelFile类提升效率
"""

import os, os.path
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from func.commonFunc import calc_spent_time
from func.sendEmail import send_email
workdir = os.getcwd()
inputDir = workdir + "\\input"
outputDir = workdir + "\\output"
cmmerge_file = inputDir + "\\merge\\cmmerge_sheet.xlsx"
outputFile = ""

inputPathFileList = []
columnName = ["名称", "IP地址IPv4", "CPUAVG", "CPUMAX", "memoryAVG", "memoryAVG"]
# [5,6,11,12,16,17]
drop_column = ["资源编号","描述","所属pod","所属科室","架构师责任人","状态","主机类型","所属宿主机IP","CPUSUM",
               "CPU采集次数","CPU采集成功次数","memorySUM","memory采集次数","memory采集成功次数"]
selectColumnLetter = ('F', 'G', 'L', 'M', 'Q', 'R')

# file1 = "D:\\github\\jsmobile\\parse_cpu_mem_excel\\input\\cmmerge1-20210913.xlsx"
# file2 = "D:\\github\\jsmobile\\parse_cpu_mem_excel\\input\\cmmerge1-20210914.xlsx"
def check_xlsx():
    xlsx_list = []
    file_date = ""
    # now_time = time.strftime("%Y%m",time.localtime())
    for f in os.listdir(inputDir):
        if f.startswith("cmmerge1-") and f.endswith(".xlsx"):
            if file_date == "":
                file_date = f[9:15]
                xlsx_list.append(f)
            elif len(file_date) == 6:
                if f[9:15] == file_date:
                    xlsx_list.append(f)
                else:
                    print(f + " : 文件日期不对！")
                    error = Exception("文件日期不对")
                    raise error
    global outputFile
    outputFile = outputDir + "\\" + "output_" + file_date + ".xlsx"
    return xlsx_list

def calc_avg(input_list) -> float:
    i_sum = 0
    if len(input_list) == 0:
        return ""
    for number in input_list:
        i_sum = i_sum + number
    result = i_sum/len(input_list)
    return "%.8f" % result

@calc_spent_time
def xlsx_merge(xlsx_list):
    writer = pd.ExcelWriter(cmmerge_file)
    sheet = 0
    for f in xlsx_list:
        daily_xlsx_file = inputDir + "\\" + f
        if daily_xlsx_file.endswith(".xlsx"):
            daily_df = pd.read_excel(daily_xlsx_file)
            #删除多余列
            daily_df.drop(columns=drop_column,axis=1,inplace=True)
            daily_df.to_excel(writer, sheet_name=str(sheet), index=False)
            sheet += 1
    writer.save()

@calc_spent_time
def xlsx_analysis():
    merge_df = pd.io.excel.ExcelFile(cmmerge_file)
    data = {}
    max_length = 0
    for sheet_name in merge_df.sheet_names:
        df = pd.read_excel(merge_df,sheet_name)
        if max_length == 0:
            max_length = len(df.index.values)
        data[int(sheet_name)] = df

    writer = pd.ExcelWriter(outputFile)
    for line in range(max_length):
        temp_list = []
        cpu_avg_list = []
        cpu_max_list = []
        memory_avg_list = []
        memory_max_list = []
        for sheet in range(len(data)):
            data[sheet].iloc[line,5]
            if len(temp_list) == 0:
                temp_list.append(data[sheet].iloc[line,0])
            if len(temp_list) == 1:
                temp_list.append(data[sheet].iloc[line,1])
            if not pd.isna(data[sheet].iloc[line,2]):
                cpu_avg_list.append(data[sheet].iloc[line,2])
            if not pd.isna(data[sheet].iloc[line,3]):
                cpu_max_list.append(data[sheet].iloc[line,3])
            if not pd.isna(data[sheet].iloc[line, 4]):
                memory_avg_list.append(data[sheet].iloc[line,4])
            if not pd.isna(data[sheet].iloc[line, 5]):
                memory_max_list.append(data[sheet].iloc[line,5])
        # 求平均
        cpu_avg = calc_avg(cpu_avg_list)
        cpu_max = calc_avg(cpu_max_list)
        memory_avg = calc_avg(memory_avg_list)
        memory_max = calc_avg(memory_max_list)
        # 添加
        temp_list.append(cpu_avg)
        temp_list.append(cpu_max)
        temp_list.append(memory_avg)
        temp_list.append(memory_max)

        # 写文件
        temp_list = np.array(temp_list).reshape(1, 6)
        df_data = pd.DataFrame(data=temp_list, columns=columnName)
        if line == 0:
            df_data.to_excel(writer, 'Sheet1', index=False)
        if line > 0:
            df_data.to_excel(writer, 'Sheet1', header=False, index=False, startrow=line + 1)
    writer.save()

@calc_spent_time
def set_style():
    wb = load_workbook(outputFile)
    ws = wb["Sheet1"]
    for i in ws.rows:
        for j in i:
            if j.value is None:
                j.fill = PatternFill('solid', fgColor='ffff00') #黄色
    ws.column_dimensions['A'].width = 13
    ws.column_dimensions['B'].width = 30
    for col in ['C','D','E','F']:
        ws.column_dimensions[col].width = 12.91

    wb.save(outputFile)

def process():
    # 表格校验
    xlsx_list = check_xlsx()
    # 多表合并
    if os.path.exists(cmmerge_file):
        os.remove(cmmerge_file)
    xlsx_merge(xlsx_list)
    # 数据分析
    xlsx_analysis()

    #添加单元格样式
    set_style()

if __name__ == "__main__":
    process()
    # send_email(outputFile)