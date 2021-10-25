"""
功能：
根据工作簿：“江苏公司IT云利用率分析-8月.xlsx”
sheet页：“B域M域x86主机信息”，“O域主机信息” 做数据分析生成 sheet页 “各业务设备汇总”
"""
from openpyxl import load_workbook
import os, os.path
import pandas as pd
import numpy as np
from func.commonFunc import trans_path
# D:\github\jsmobile\parse_cpu_mem_excel\input\江苏公司IT云利用率分析-8月.xlsx
# ori_file = trans_path(input("请输入要分析的xlsx表格的完整路径："))
ori_file = trans_path("D:\github\jsmobile\parse_cpu_mem_excel\input\江苏公司IT云利用率分析-8月.xlsx")
header_volumn = ["业务系统","所属科室","架构室维护责任人","主机总数","物理机数量","虚拟机数量","本月上线","本月下线","本月cpu平均利用率","本月内存平均利用率"]
sheet_B_M_area = "B域M域x86主机信息"
sheet_O_area = "O域主机信息"
sheet_onoff_line = ori_file[-7:-5] + "上下线主机信息"
list_b_m_business_system = ["集客调度中心功能","备份/管理节点/数据库等","采集/结算","CRM-bes并行环境","测试环境",
"广告资源系统","电子渠道","订单、认证中心","4A/SMP","CRM-C30","CRM-ICRM","宝兰德Webgate",
"车联网","CRM-性能环境","CRM-C00","4A/安全","BOMC","BOSS","阿波罗渠道","CRM-携号转网",
"订单中心","订单认证系统","大数据对外","BOSS模拟出账","CRM","地市统一支撑","CRM工具",
"阿拉盯","CRM-应急","DMS","经营分析","能力开放平台","渠道系统扩容工程","认证中心",
"统一PAAS平台","物联网","写卡系统","无人认领","信息化","一级BOSS","一级产商品中心","营改增",
"增值税税控发票管理系统","招聘考核系统","中移在线","终端管家","卓望SIMS/局数据系统"]
list_o_business_system = ["O域网管云","O域信令"]
#汇总信息
summary_list = [["B域M域主机信息汇总","","",0,0,0,"","",0,0],
                ["8月主机信息汇总","","",0,0,0,"","",0,0]]

l_b_m_cpu_avg = []
l_b_m_mem_avg = []
l_all_cpu_avg = []
l_all_mem_avg = []

ERROR_ORI_FILE_NOT_FOUND = 127
ERROR_HOSTS_COUNT = 128
ERROR_ORI_FILENAME = 129
def check():
    """
    :表格校验模块
    :return:
    """
    # 检查xlsx文件是否存在
    if os.path.exists(ori_file):
        pass
    else:
        raise ERROR_ORI_FILE_NOT_FOUND
    # 检查文件名格式是否正确
    if ori_file[-6:] != "月.xlsx":
        raise ERROR_ORI_FILENAME

def initial(sheetName,listBusinessSystem):
    """
    :数据清洗模块
    :功能 :将总表数据清洗，按业务系统分类，存入dict中
    :return:  dict_temp
    """
    df_sheet = pd.read_excel(ori_file,sheet_name=sheetName)

    dict_temp = {}
    # 初始化list_temp
    list_temp = []
    for i in range(len(listBusinessSystem)):
        list_temp.append([])

    # if sheetName == sheet_B_M_area:
    #     summary_list[0][3] = len(df_sheet.index.values)
    # if sheetName == sheet_O_area:
    #     summary_list[1][3] = summary_list[0][3] + len(df_sheet.index.values)

    # 按业务系统分类，生成各个列表
    for line_num in range(len(df_sheet.index.values)):
        if sheetName == sheet_O_area:
            ### begin : 此处顺带做一下cpu和mem的汇总 ###
            all_cpu_avg = df_sheet.iloc[line_num, 11]
            all_mem_avg = df_sheet.iloc[line_num, 12]
            if type(all_cpu_avg) != str:
                l_all_cpu_avg.append(all_cpu_avg)
            if type(all_mem_avg) != str:
                l_all_mem_avg.append(all_mem_avg)
            ### end : 此处顺带做一下cpu和mem的汇总 ###
            # 处理O域sheet页
            if "网管云" in df_sheet.iloc[line_num, 5]:
                system_name = "O域网管云"
            elif "信令" in df_sheet.iloc[line_num, 5]:
                system_name = "O域信令"
        else:
            ### begin : 此处顺带做一下cpu和mem的汇总 ###
            b_m_cpu_avg = df_sheet.iloc[line_num, 11]
            b_m_mem_avg = df_sheet.iloc[line_num, 12]
            all_cpu_avg = df_sheet.iloc[line_num, 11]
            all_mem_avg = df_sheet.iloc[line_num, 12]
            if type(b_m_cpu_avg) != str:
                l_b_m_cpu_avg.append(b_m_cpu_avg)
            if type(b_m_mem_avg) != str:
                l_b_m_mem_avg.append(b_m_mem_avg)
            if type(all_cpu_avg) != str:
                l_all_cpu_avg.append(all_cpu_avg)
            if type(all_mem_avg) != str:
                l_all_mem_avg.append(all_mem_avg)
            ### end   : 此处顺带做一下cpu和mem的汇总 ###
            # 处理B域M域sheet页
            system_name = df_sheet.iloc[line_num, 7]

        system_index = listBusinessSystem.index(system_name)
        list_temp[system_index].append(df_sheet.iloc[line_num].values)

    ### begin : 此处顺带做一下cpu和mem的汇总 ###
    if sheetName == sheet_O_area:
        summary_list[0][8] = calc_avg(l_b_m_cpu_avg)
        summary_list[0][9] = calc_avg(l_b_m_mem_avg)
        summary_list[1][8] = calc_avg(l_all_cpu_avg)
        summary_list[1][9] = calc_avg(l_all_mem_avg)
    ### begin : 此处顺带做一下cpu和mem的汇总 ###

    # 将多个列表转成numpy数组放入字典
    a = 0
    for j in list_temp:
        if len(j) != 0:
            dict_temp[listBusinessSystem[a]] = np.array(j)
        a = a + 1

    return dict_temp

def initial_onoff_sheet(sheetName):
    df = pd.read_excel(ori_file, sheet_name=sheetName)
    p_online,p_offline,v_online,v_offline,all_online,all_offline = 0,0,0,0,0,0
    for line_num in range(len(df.index.values)):
        if df.iloc[line_num, 3] == "物理机":
            if df.iloc[line_num, 4] == "上线":
                p_online = p_online + 1
            elif df.iloc[line_num, 4] == "下线":
                p_offline = p_offline + 1
        elif df.iloc[line_num, 3] == "虚拟机":
            if df.iloc[line_num, 4] == "上线":
                v_online = v_online + 1
            elif df.iloc[line_num, 4] == "下线":
                v_offline = v_offline + 1
    all_online = p_online + v_online
    all_offline = v_online + v_offline
    for business in np.unique(df["所属业务"].values):
        #if df.iloc[]
        pass

def get_department(lis):
    result = ""
    for dep in np.unique(np.array(lis)[...,8]):
        result = result + str(dep) + '/'
    return result[0:-1]

def get_maintenance_person(lis):
    result = ""
    for per in np.unique(np.array(lis)[...,9]):
        result = result + str(per) + '/'
    return result[0:-1]

def get_hosts_number(lis):
    x,y,z = 0,0,0
    # 计算主机总数
    x = len(lis)
    # 计算 物理机y 和 虚拟机z 总数
    for line in lis:
        if line[3] == "普通服务器":
            y = y + 1
        elif line[3] == "VMware":
            z = z + 1
    if y + z == x:
        return x, y, z
    else:
        raise ERROR_HOSTS_COUNT

def calc_avg(input_list) -> float:
    i_sum = 0
    if len(input_list) == 0:
        return ""
    for number in input_list:
        i_sum = i_sum + number
    result = i_sum/len(input_list)
    return "%.2f%%" % (result * 100)

def get_avg_rate(lis):
    l_cpu = []
    l_mem = []
    for line in lis:
        if type(line[11]) != str:
            l_cpu.append(line[11])
        if type(line[12]) != str:
            l_mem.append(line[12])
    return calc_avg(l_cpu), calc_avg(l_mem)

def analysis(dic):
    """
    数据分析模块
    :param dic:
    :return:
    """
    l = []
    for key in dic.keys():
        line_list = []
        # 初始化各列
        dict_value = dic[key]
        business_system = key
        department = get_department(dict_value)
        maintenance_person = get_maintenance_person(dict_value)
        hosts_number,physical_number,virtual_number = get_hosts_number(dict_value)
        online_hosts = ""
        offline_hosts = ""
        cpu_avg_rate,mem_avg_rate = get_avg_rate(dict_value)
        # 合并
        for i in [business_system,department,maintenance_person,hosts_number,physical_number,virtual_number,
                  online_hosts,offline_hosts,cpu_avg_rate,mem_avg_rate]:
            line_list.append(i)
        l.append(line_list)
    return l

def writer_sheet(l):
    df = pd.DataFrame(l, columns=header_volumn)
    # 写文件
    book = load_workbook(ori_file)
    writer = pd.ExcelWriter(ori_file)
    writer.book = book
    df.to_excel(writer, sheet_name="各业务设备汇总auto", index=False)
    writer.save()

def do_summary(l):
    i = 0
    while i < len(l):
        if i < len(l)-2:
            summary_list[0][3] = summary_list[0][3] + l[i][3]
            summary_list[0][4] = summary_list[0][4] + l[i][4]
            summary_list[0][5] = summary_list[0][5] + l[i][5]
            if i == len(l)-3:
                summary_list[1][3] = summary_list[0][3]
                summary_list[1][4] = summary_list[0][4]
                summary_list[1][5] = summary_list[0][5]
        else:
            summary_list[1][3] = summary_list[1][3] + l[i][3]
            summary_list[1][4] = summary_list[1][4] + l[i][4]
            summary_list[1][5] = summary_list[1][5] + l[i][5]
        i+=1

def set_style():
    wb = load_workbook(ori_file)
    ws = wb["各业务设备汇总auto"]
    ws.column_dimensions['A'].width = 28.55
    for col in ['B','C','D','E','F','G','H']:
        ws.column_dimensions[col].width = 11
    ws.column_dimensions['I'].width = 17.27
    ws.column_dimensions['J'].width = 17.27
    wb.save(ori_file)

if __name__ == '__main__':
    # 合规校验
    check()

    dict_result = initial(sheet_B_M_area,list_b_m_business_system)
    dict_O = initial(sheet_O_area,list_o_business_system)
    dict_result.update(dict_O)
    merge_list = analysis(dict_result)

    initial_onoff_sheet(sheet_onoff_line)

    do_summary(merge_list)

    # 将各业务系统的统计列表 和 两行汇总列表合并
    for row in summary_list:
        merge_list.append(row)

    writer_sheet(merge_list)

    set_style()